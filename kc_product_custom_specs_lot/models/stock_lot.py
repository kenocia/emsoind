# -*- coding: utf-8 -*-

import csv
import io

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools.float_utils import float_is_zero

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class StockLot(models.Model):
    _inherit = 'stock.lot'

    technical_description = fields.Text(string='Descripción técnica')
    technical_key = fields.Char(string='Clave técnica', index=True)
    sale_order_id = fields.Many2one('sale.order', string='Pedido de venta', ondelete='set null')
    sale_order_line_id = fields.Many2one(
        'sale.order.line',
        string='Línea de venta',
        ondelete='set null',
    )
    partner_id = fields.Many2one('res.partner', string='Cliente')
    source_type = fields.Selection(
        selection=[
            ('sale', 'Venta'),
            ('inventory', 'Inventario'),
            ('manual', 'Manual'),
        ],
        string='Origen',
        default='manual',
    )
    kc_authorized_employee_id = fields.Many2one(
        'hr.employee',
        string='Autorizado por',
        ondelete='set null',
        copy=False,
    )
    technical_value_ids = fields.One2many(
        'custom.stock.lot.technical.value',
        'lot_id',
        string='Especificaciones técnicas',
    )
    kc_technical_configuration_id = fields.Many2one(
        'product.technical.configuration',
        string='Configuración técnica (matriz)',
        compute='_compute_kc_technical_configuration_id',
        store=True,
        readonly=True,
    )
    kc_config_standard_price = fields.Float(
        string='Costo catálogo (matriz)',
        related='kc_technical_configuration_id.standard_price',
        readonly=True,
    )

    @api.depends('product_id', 'technical_key')
    def _compute_kc_technical_configuration_id(self):
        Configuration = self.env['product.technical.configuration']
        for lot in self:
            if lot.product_id and lot.technical_key:
                lot.kc_technical_configuration_id = Configuration.find_by_product_and_key(
                    lot.product_id, lot.technical_key,
                )
            else:
                lot.kc_technical_configuration_id = False

    @api.model
    def _kc_resolve_unit_cost(self, product, technical_key=None, lot=None, company=None):
        """Costo unitario operativo: lote → matriz técnica → producto genérico."""
        company = company or self.env.company
        if lot and product and product.lot_valuated:
            lot_cost = lot.with_company(company).standard_price
            if lot_cost:
                return lot_cost
        return self._kc_get_standard_price_from_config(product, technical_key)

    def _update_technical_meta_from_values(self):
        mixin = self.env['custom.technical.value.mixin']
        for lot in self:
            if not lot.technical_value_ids:
                continue
            lot.technical_description = mixin.build_technical_description(
                lot.technical_value_ids
            )
            lot.technical_key = mixin.build_technical_key(lot.technical_value_ids)

    @api.model
    def _kc_get_standard_price_from_config(self, product, technical_key, configuration=None):
        """Resuelve el costo inicial de un lote nuevo desde la matriz técnica."""
        if configuration:
            return configuration._kc_get_standard_price_for_lot(product=product)
        if product and technical_key:
            configuration = self.env['product.technical.configuration'].find_by_product_and_key(
                product, technical_key,
            )
            if configuration:
                return configuration._kc_get_standard_price_for_lot(product=product)
        if product:
            return product.with_company(self.env.company).standard_price
        return 0.0

    def _kc_prepare_lot_cost_vals(self, product, technical_key, configuration=None):
        cost = self._kc_get_standard_price_from_config(
            product, technical_key, configuration=configuration,
        )
        if not product.lot_valuated or not cost:
            return {}
        return {'standard_price': cost}

    @api.model
    def _get_or_create_for_sale_line(self, sale_line, authorized_employee=None):
        """Reutiliza un lote de producción de la misma OV; si no existe, lo crea.

        No reutiliza lotes de inventario (source_type=inventory) para no
        mezclar producción dentro de existencias históricas.

        También reutiliza lotes de venta huérfanos (sin stock / sin kc_entry_id)
        vinculados a la misma OV: el nombre de lote es determinístico y único
        por empresa+producto.
        """
        if sale_line.technical_key and sale_line.order_id:
            order_id = sale_line.order_id.id
            domain = [
                ('product_id', '=', sale_line.product_id.id),
                ('technical_key', '=', sale_line.technical_key),
                ('source_type', '!=', 'inventory'),
                '|',
                ('kc_sale_order_id', '=', order_id),
                ('sale_order_id', '=', order_id),
            ]
            existing = self.search(domain, limit=1)
            if existing:
                return existing
        return self._create_from_sale_line(
            sale_line, authorized_employee=authorized_employee,
        )

    @api.model
    def _create_from_sale_line(self, sale_line, authorized_employee=None):
        mixin = self.env['custom.technical.value.mixin']
        product = sale_line.product_id
        company_id = sale_line.company_id.id
        name = mixin.build_lot_name(
            product.default_code or product.name,
            sale_line.technical_value_ids,
            order_name=sale_line.order_id.name,
        )
        # Seguridad ante constraint unique(product, name, company): reutilizar
        # si el lote ya existe (p. ej. creado en un intento previo sin stock).
        existing = self.search([
            ('name', '=', name),
            ('product_id', '=', product.id),
            '|',
            ('company_id', '=', company_id),
            ('company_id', '=', False),
        ], limit=1)
        if existing:
            vals = {}
            if not existing.sale_order_id:
                vals['sale_order_id'] = sale_line.order_id.id
            if not existing.sale_order_line_id:
                vals['sale_order_line_id'] = sale_line.id
            if (
                'kc_sale_order_id' in existing._fields
                and not existing.kc_sale_order_id
            ):
                vals['kc_sale_order_id'] = sale_line.order_id.id
            if not existing.technical_key and sale_line.technical_key:
                vals['technical_key'] = sale_line.technical_key
            if existing.source_type != 'sale':
                vals['source_type'] = 'sale'
            if authorized_employee and not existing.kc_authorized_employee_id:
                vals['kc_authorized_employee_id'] = authorized_employee.id
            if vals:
                existing.write(vals)
            return existing
        lot = self.create({
            'name': name,
            'product_id': product.id,
            'company_id': company_id,
            'technical_description': sale_line.technical_description,
            'technical_key': sale_line.technical_key,
            'sale_order_id': sale_line.order_id.id,
            'sale_order_line_id': sale_line.id,
            'partner_id': sale_line.order_id.partner_id.id,
            'source_type': 'sale',
            'kc_authorized_employee_id': authorized_employee.id if authorized_employee else False,
            **self._kc_prepare_lot_cost_vals(product, sale_line.technical_key),
        })
        lot._copy_technical_values_from_sale_line(sale_line)
        return lot

    def _copy_technical_values_from_sale_line(self, sale_line):
        self.ensure_one()
        commands = []
        for tv in sale_line.technical_value_ids.sorted('sequence'):
            commands.append((0, 0, {
                'attribute_id': tv.attribute_id.id,
                'value_text': tv.value_text,
                'value_number': tv.value_number,
                'value_id': tv.value_id.id,
                'uom_id': tv.uom_id.id,
                'sequence': tv.sequence,
            }))
        self.write({'technical_value_ids': commands})
        self._update_technical_meta_from_values()

    @api.model
    def _create_lot_from_configuration(self, configuration, product=None,
                                       company=None):
        """Crea un lote técnico a partir de una configuración de producto
        (matriz `product.technical.configuration`), SIN vincularlo a una orden
        de venta. Pensado para abastecimiento/producción: el nombre usa fecha +
        correlativo en lugar del pedido. Siempre crea un lote nuevo (cada
        producción es una entrada física distinta que comparte technical_key)."""
        if not configuration:
            return self.browse()
        mixin = self.env['custom.technical.value.mixin']
        if not product:
            product = configuration.product_tmpl_id.product_variant_id
        if not company:
            company = self.env.company
        date_str = fields.Date.context_today(self).strftime('%Y%m%d')
        sequence = self.env['ir.sequence'].next_by_code('stock.lot.technical') or '0001'
        name = mixin.build_lot_name(
            product.default_code or product.name,
            configuration.technical_value_ids,
            date_str=date_str,
            sequence=sequence,
        )
        lot = self.create({
            'name': name,
            'product_id': product.id,
            'company_id': company.id,
            'technical_description': configuration.technical_description,
            'technical_key': configuration.technical_key,
            'source_type': 'inventory',
            **self._kc_prepare_lot_cost_vals(
                product, configuration.technical_key, configuration=configuration,
            ),
        })
        lot._copy_technical_values_from_configuration(configuration)
        return lot

    def _copy_technical_values_from_configuration(self, configuration):
        self.ensure_one()
        commands = []
        for tv in configuration.technical_value_ids.sorted('sequence'):
            commands.append((0, 0, {
                'attribute_id': tv.attribute_id.id,
                'value_text': tv.value_text,
                'value_number': tv.value_number,
                'value_id': tv.value_id.id,
                'uom_id': tv.uom_id.id,
                'sequence': tv.sequence,
            }))
        self.write({'technical_value_ids': commands})
        self._update_technical_meta_from_values()

    # -------------------------------------------------------------------------
    # Inventario inicial por configuración técnica (lote)
    # -------------------------------------------------------------------------

    @api.model
    def _kc_inventory_import_header_aliases(self):
        return {
            'legacy_sku': {
                'legacy_sku', 'sku', 'sku_original', 'skuoriginal',
                'sku_referencia', 'skureferencia', 'referencia', 'codigo',
                'codigo_sku', 'codigosku', 'sku_referencia_actual',
            },
            'product_default_code': {
                'product_default_code', 'codigo_base', 'codigobase',
                'codigo_producto', 'codigoproducto', 'producto_base',
                'productobase', 'codigo_base_sugerido',
            },
            'configuration_key': {
                'configuration_key', 'clave_configuracion', 'claveconfiguracion',
                'clave_de_configuracion',
            },
            'technical_key': {
                'technical_key', 'clave_tecnica', 'clavetecnica',
            },
            'quantity': {
                'quantity', 'cantidad', 'existencia', 'existencias', 'qty',
                'cant', 'stock', 'inventario',
            },
            'lot_name': {
                'lot_name', 'lote', 'lote_sugerido', 'lotesugerido',
                'numero_lote', 'numerolote', 'name',
            },
            'unit_cost': {
                'unit_cost', 'costo', 'costo_unitario', 'costounitario',
                'standard_price', 'costo_estandar',
            },
            'location_name': {
                'location_name', 'ubicacion', 'location', 'bodega',
            },
        }

    @api.model
    def _kc_normalize_import_header(self, header):
        mixin = self.env['custom.technical.value.mixin']
        text = mixin._normalize_token(header or '')
        aliases = self._kc_inventory_import_header_aliases()
        for canonical, names in aliases.items():
            if text in names:
                return canonical
        return text

    @api.model
    def _kc_parse_inventory_quantity(self, raw_value):
        if raw_value is None:
            return 0.0
        text = str(raw_value).strip()
        if not text:
            return 0.0
        text = text.replace(',', '.')
        try:
            return float(text)
        except ValueError as exc:
            raise UserError(
                _('La cantidad "%(value)s" no es un número válido.') % {'value': raw_value}
            ) from exc

    @api.model
    def _kc_parse_inventory_cost(self, raw_value):
        if raw_value is None:
            return 0.0
        text = str(raw_value).strip()
        if not text:
            return 0.0
        text = text.replace(',', '.')
        try:
            return float(text)
        except ValueError as exc:
            raise UserError(
                _('El costo "%(value)s" no es un número válido.') % {'value': raw_value}
            ) from exc

    @api.model
    def _kc_normalize_import_row(self, row):
        normalized = {}
        for key, value in row.items():
            canonical = self._kc_normalize_import_header(key)
            if not canonical:
                continue
            if value is None:
                normalized[canonical] = ''
            elif isinstance(value, float) and canonical in ('legacy_sku', 'product_default_code'):
                normalized[canonical] = str(int(value)) if value == int(value) else str(value)
            else:
                normalized[canonical] = str(value).strip() if value is not False else ''
        return normalized

    @api.model
    def _kc_find_configuration_for_inventory_import(self, row):
        Configuration = self.env['product.technical.configuration']
        configuration_key = (row.get('configuration_key') or '').strip()
        if configuration_key:
            config = Configuration.search([
                ('configuration_key', '=', configuration_key),
                ('active', '=', True),
            ], limit=1)
            if config:
                return config

        legacy_sku = (row.get('legacy_sku') or '').strip()
        if legacy_sku:
            config = Configuration.search([
                ('legacy_sku', '=', legacy_sku),
                ('active', '=', True),
            ], limit=1)
            if config:
                return config

        product_code = (row.get('product_default_code') or '').strip()
        technical_key = (row.get('technical_key') or '').strip()
        if product_code and technical_key:
            tmpl = self.env['product.template'].search([
                ('default_code', '=', product_code),
            ], limit=1)
            if tmpl:
                return Configuration.search([
                    ('product_tmpl_id', '=', tmpl.id),
                    ('technical_key', '=', technical_key),
                    ('active', '=', True),
                ], limit=1)

        if technical_key and not product_code:
            configs = Configuration.search([
                ('technical_key', '=', technical_key),
                ('active', '=', True),
            ])
            if len(configs) == 1:
                return configs

        return Configuration.browse()

    @api.model
    def _kc_resolve_inventory_location(self, location, location_name=None, company=None):
        company = company or self.env.company
        if location:
            return location
        if location_name:
            loc = self.env['stock.location'].search([
                ('complete_name', 'ilike', location_name),
                ('usage', '=', 'internal'),
                ('company_id', 'in', [company.id, False]),
            ], limit=1)
            if loc:
                return loc
        loc = self.env['stock.location'].search([
            ('complete_name', 'ilike', 'Bodega PT'),
            ('usage', '=', 'internal'),
            ('company_id', 'in', [company.id, False]),
        ], limit=1)
        return loc

    @api.model
    def _get_or_create_inventory_lot(self, configuration, company, lot_name=None,
                                     inventory_date=None, unit_cost=None):
        """Obtiene o crea un lote de inventario para una configuración técnica.

        :return: tuple (stock.lot, created: bool)
        """
        product = configuration.product_tmpl_id.product_variant_id
        existing = self.search([
            ('product_id', '=', product.id),
            ('technical_key', '=', configuration.technical_key),
            ('source_type', '=', 'inventory'),
            ('company_id', '=', company.id),
        ], limit=1)
        if existing:
            if unit_cost and product.lot_valuated:
                existing.with_company(company).standard_price = unit_cost
            return existing, False

        if lot_name:
            lot_vals = {
                'name': lot_name,
                'product_id': product.id,
                'company_id': company.id,
                'technical_description': configuration.technical_description,
                'technical_key': configuration.technical_key,
                'source_type': 'inventory',
            }
            cost = unit_cost or self._kc_get_standard_price_from_config(
                product, configuration.technical_key, configuration=configuration,
            )
            if product.lot_valuated and cost:
                lot_vals['standard_price'] = cost
            lot = self.create(lot_vals)
            lot._copy_technical_values_from_configuration(configuration)
            return lot, True

        inv_date = inventory_date or fields.Date.context_today(self)
        if isinstance(inv_date, str):
            inv_date = fields.Date.from_string(inv_date)
        date_str = inv_date.strftime('%Y%m%d')
        sequence = self.env['ir.sequence'].next_by_code('stock.lot.technical') or '0001'
        mixin = self.env['custom.technical.value.mixin']
        name = mixin.build_lot_name(
            product.default_code or product.name,
            configuration.technical_value_ids,
            date_str=date_str,
            sequence=sequence,
        )
        lot = self.create({
            'name': name,
            'product_id': product.id,
            'company_id': company.id,
            'technical_description': configuration.technical_description,
            'technical_key': configuration.technical_key,
            'source_type': 'inventory',
            **self._kc_prepare_lot_cost_vals(
                product,
                configuration.technical_key,
                configuration=configuration,
            ),
        })
        if unit_cost and product.lot_valuated:
            lot.with_company(company).standard_price = unit_cost
        lot._copy_technical_values_from_configuration(configuration)
        return lot, True

    @api.model
    def _kc_read_inventory_import_rows(self, file_content, file_name=None):
        name = (file_name or '').lower()
        if name.endswith('.xlsx') or name.endswith('.xlsm'):
            return self._kc_read_inventory_xlsx(file_content)
        if name.endswith('.csv'):
            return self._kc_read_inventory_csv(file_content)
        if file_content[:2] == b'PK':
            return self._kc_read_inventory_xlsx(file_content)
        return self._kc_read_inventory_csv(file_content)

    @api.model
    def _kc_read_inventory_csv(self, file_content):
        try:
            text = file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_content.decode('latin-1')
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise UserError(_('El archivo está vacío o no tiene encabezados.'))
        return [self._kc_normalize_import_row(row) for row in reader]

    @api.model
    def _kc_read_inventory_xlsx(self, file_content):
        if not load_workbook:
            raise UserError(_(
                'Para importar Excel (.xlsx) instale la librería openpyxl en el servidor.'
            ))
        book = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
        sheet = book.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration as exc:
            raise UserError(_('El archivo Excel está vacío.')) from exc
        header_map = {
            idx: self._kc_normalize_import_header(header)
            for idx, header in enumerate(headers)
            if header is not None and str(header).strip()
        }
        if not header_map:
            raise UserError(_('No se encontraron encabezados válidos en el Excel.'))
        rows = []
        for raw_row in rows_iter:
            if not raw_row or not any(
                cell is not None and str(cell).strip() for cell in raw_row
            ):
                continue
            row = {}
            for idx, canonical in header_map.items():
                if idx < len(raw_row):
                    row[canonical] = raw_row[idx]
            rows.append(self._kc_normalize_import_row(row))
        book.close()
        return rows

    @api.model
    def import_initial_inventory(
        self,
        file_content,
        file_name=None,
        location=None,
        inventory_date=None,
        default_location_name=None,
        merge_existing_lot=True,
    ):
        """Carga inventario inicial: configuración técnica → lote → quant."""
        rows = self._kc_read_inventory_import_rows(file_content, file_name=file_name)
        if not rows:
            raise UserError(_('El archivo no contiene filas de datos.'))

        company = self.env.company
        location = self._kc_resolve_inventory_location(
            location,
            location_name=default_location_name,
            company=company,
        )
        if not location:
            raise UserError(_(
                'No se encontró la ubicación destino. Indique la ubicación en el '
                'asistente (ej. ESI/Bodega PT) o incluya la columna "Ubicación".'
            ))

        if inventory_date:
            if isinstance(inventory_date, str):
                inventory_date = fields.Date.from_string(inventory_date)
        else:
            inventory_date = fields.Date.context_today(self)

        Quant = self.env['stock.quant'].sudo()
        lots_created_ids = set()
        lots_reused_ids = set()
        applied_rows = 0
        skipped = 0
        errors = []

        for index, row in enumerate(rows, start=2):
            qty = self._kc_parse_inventory_quantity(row.get('quantity'))
            if float_is_zero(qty, precision_rounding=0.0001):
                skipped += 1
                continue

            row_location = self._kc_resolve_inventory_location(
                location,
                location_name=(row.get('location_name') or '').strip() or None,
                company=company,
            )
            if not row_location:
                skipped += 1
                errors.append(_(
                    'Fila %(row)s: ubicación no encontrada (%(loc)s).'
                ) % {'row': index, 'loc': row.get('location_name') or ''})
                continue

            config = self._kc_find_configuration_for_inventory_import(row)
            if not config:
                skipped += 1
                errors.append(_(
                    'Fila %(row)s: configuración no encontrada (SKU=%(sku)s, clave=%(key)s).'
                ) % {
                    'row': index,
                    'sku': row.get('legacy_sku') or '',
                    'key': row.get('configuration_key') or row.get('technical_key') or '',
                })
                continue

            product = config.product_tmpl_id.product_variant_id
            if product.tracking != 'lot':
                skipped += 1
                errors.append(_(
                    'Fila %(row)s: el producto %(product)s no está configurado con '
                    'seguimiento por lote.'
                ) % {'row': index, 'product': product.display_name})
                continue

            unit_cost = self._kc_parse_inventory_cost(row.get('unit_cost'))
            lot_name = (row.get('lot_name') or '').strip() or None
            lot, lot_created = self._get_or_create_inventory_lot(
                config,
                company,
                lot_name=lot_name,
                inventory_date=inventory_date,
                unit_cost=unit_cost or None,
            )
            if lot_created:
                lots_created_ids.add(lot.id)
            else:
                lots_reused_ids.add(lot.id)

            existing_quant = Quant.search([
                ('product_id', '=', product.id),
                ('location_id', '=', row_location.id),
                ('lot_id', '=', lot.id),
                ('company_id', '=', company.id),
            ], limit=1)

            if existing_quant and not merge_existing_lot:
                skipped += 1
                errors.append(_(
                    'Fila %(row)s: ya existe stock para lote %(lot)s en %(loc)s.'
                ) % {
                    'row': index,
                    'lot': lot.name,
                    'loc': row_location.display_name,
                })
                continue

            inv_ctx = {
                'inventory_mode': True,
                'inventory_name': _('Inventario inicial PT %(date)s') % {
                    'date': fields.Date.to_string(inventory_date or fields.Date.context_today(self)),
                },
            }
            if inventory_date:
                inv_ctx['force_period_date'] = fields.Date.to_string(inventory_date)

            if existing_quant:
                target_qty = existing_quant.quantity + qty
                existing_quant.with_context(inv_ctx).write({
                    'inventory_quantity': target_qty,
                })
                existing_quant.with_context(inv_ctx)._apply_inventory()
            else:
                new_quant = Quant.with_context(inv_ctx).create({
                    'product_id': product.id,
                    'location_id': row_location.id,
                    'lot_id': lot.id,
                    'inventory_quantity': qty,
                })
                new_quant.with_context(inv_ctx)._apply_inventory()
            applied_rows += 1

        created_lots = len(lots_created_ids)
        updated_lots = len(lots_reused_ids)
        message = _(
            '%(rows)s filas aplicadas, %(created)s lotes creados, '
            '%(updated)s lotes reutilizados, %(skipped)s omitidas.'
        ) % {
            'rows': applied_rows,
            'created': created_lots,
            'updated': updated_lots,
            'skipped': skipped,
        }
        return {
            'applied_rows': applied_rows,
            'created_lots': created_lots,
            'updated_lots': updated_lots,
            'skipped': skipped,
            'errors': errors[:30],
            'message': message,
            'location': location.display_name,
        }
